import json
import base64
import io
import os
from http.server import BaseHTTPRequestHandler

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage


class handler(BaseHTTPRequestHandler):

    def do_OPTIONS(self):
        self.send_response(200)
        self._cors()
        self.end_headers()

    def do_POST(self):
        length = int(self.headers.get("Content-Length", 0))
        body   = json.loads(self.rfile.read(length) or b"{}")
        header  = body.get("header", {})
        entries = body.get("entries", [])

        try:
            # Template ada di public/templates/ — Vercel cwd = project root
            template_path = os.path.join(os.getcwd(), "public", "templates", "scenario-test-template.xlsx")
            if not os.path.exists(template_path):
                self._error(f"Template tidak ditemukan: {template_path}")
                return

            wb = load_workbook(template_path)
            ws = wb.active

            # Sheet name
            safe = (header.get("judulDokumen") or "Sheet1").strip()[:31]
            ws.title = safe

            # Header fields
            ws["D4"] = header.get("keterangan", "")
            ws["D5"] = header.get("aplikasi", "")
            ws["D6"] = header.get("modul", "")
            ws["D7"] = header.get("createdBy", "")
            ws["D8"] = header.get("testedBy", "")
            ws["D9"] = header.get("targetFinish", "")
            ws["B12"] = header.get("judulDokumen", "")
            ws["B12"].font = Font(bold=True, size=13, name="Calibri")
            ws["B12"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

            # Constants
            ROW_H_PT = 14.4
            COL_D_PX = 788
            ROW_H_PX = ROW_H_PT * 96 / 72

            thin = Side(style="thin", color="000000")

            def lr(cell):
                cell.border = Border(left=thin, right=thin)

            def lr_bottom(cell):
                cell.border = Border(left=thin, right=thin, bottom=thin)

            def f11(cell, bold=False, align="general", color="000000"):
                cell.font = Font(bold=bold, size=11, name="Calibri", color=color)
                cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)

            def status_font(s):
                if s == "NOT OK":
                    return Font(bold=True, size=11, name="Calibri", color="FF0000")
                return Font(bold=(s == "OK"), size=11, name="Calibri", color="000000")

            def embed_image(data_url, row):
                if not data_url or "base64," not in data_url:
                    return row
                try:
                    b64 = data_url.split("base64,")[1]
                    pil = PILImage.open(io.BytesIO(base64.b64decode(b64)))
                    ow, oh = pil.size
                    dh = int(oh * COL_D_PX / ow)
                    nr = max(1, int(dh / ROW_H_PX) + 1)
                    for r in range(row, row + nr):
                        ws.row_dimensions[r].height = ROW_H_PT
                    buf = io.BytesIO()
                    pil.save(buf, format="PNG")
                    buf.seek(0)
                    xl = XLImage(buf)
                    xl.width = COL_D_PX
                    xl.height = dh
                    ws.add_image(xl, f"D{row}")
                    return row + nr
                except Exception:
                    return row + 1

            current_row = 26

            for entry in entries:
                status = entry.get("status", "")
                tgl    = entry.get("tanggalTest", "")

                ws.row_dimensions[current_row].height = ROW_H_PT
                ws[f"B{current_row}"] = entry.get("no", "")
                ws[f"C{current_row}"] = entry.get("object", "")
                ws[f"D{current_row}"] = entry.get("keterangan", "")
                ws[f"E{current_row}"] = tgl
                ws[f"F{current_row}"] = status

                f11(ws[f"B{current_row}"], align="center")
                f11(ws[f"C{current_row}"], bold=True)
                ws[f"D{current_row}"].font = Font(size=11, name="Calibri")
                ws[f"D{current_row}"].alignment = Alignment(wrap_text=True, vertical="center")
                f11(ws[f"E{current_row}"], align="center")
                ws[f"F{current_row}"].font = status_font(status)
                ws[f"F{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

                for col in ["B", "C", "D", "E", "F"]:
                    lr(ws[f"{col}{current_row}"])

                current_row += 1

                for img in entry.get("images", []):
                    current_row = embed_image(img.get("dataUrl", ""), current_row)

                for ss in entry.get("subSections", []):
                    if ss.get("deskripsi"):
                        ws.row_dimensions[current_row].height = ROW_H_PT
                        ws[f"D{current_row}"] = ss["deskripsi"]
                        ws[f"D{current_row}"].font = Font(size=11, name="Calibri")
                        ws[f"D{current_row}"].alignment = Alignment(wrap_text=True, vertical="center")
                        for col in ["B", "C", "D", "E", "F"]:
                            lr(ws[f"{col}{current_row}"])
                        current_row += 1
                    for img in ss.get("images", []):
                        current_row = embed_image(img.get("dataUrl", ""), current_row)

                if entry.get("subKeterangan"):
                    ws.row_dimensions[current_row].height = ROW_H_PT
                    ws[f"D{current_row}"] = entry["subKeterangan"]
                    ws[f"E{current_row}"] = tgl
                    ws[f"D{current_row}"].font = Font(size=11, name="Calibri")
                    ws[f"D{current_row}"].alignment = Alignment(vertical="center")
                    ws[f"E{current_row}"].font = Font(size=11, name="Calibri")
                    ws[f"E{current_row}"].alignment = Alignment(horizontal="center", vertical="center")
                    for col in ["B", "C", "D", "E", "F"]:
                        lr(ws[f"{col}{current_row}"])
                    current_row += 1

                for col in ["B", "C", "D", "E", "F"]:
                    lr_bottom(ws[f"{col}{current_row - 1}"])

            # Save to bytes
            out = io.BytesIO()
            wb.save(out)
            out.seek(0)
            xlsx_bytes = out.read()

            fname = (header.get("judulDokumen") or "scenario-test").strip()
            for ch in r'\/:*?"<>|':
                fname = fname.replace(ch, "_")

            self.send_response(200)
            self._cors()
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", f'attachment; filename="{fname}.xlsx"')
            self.send_header("Content-Length", str(len(xlsx_bytes)))
            self.end_headers()
            self.wfile.write(xlsx_bytes)

        except Exception as e:
            self._error(str(e))

    def _cors(self):
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.send_header("Access-Control-Allow-Methods", "POST, OPTIONS")

    def _error(self, msg):
        body = json.dumps({"error": msg}).encode()
        self.send_response(500)
        self._cors()
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)
