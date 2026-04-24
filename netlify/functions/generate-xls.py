import json
import base64
import io
import os
import tempfile

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage


def handler(event, context):
    # Handle CORS preflight
    if event.get("httpMethod") == "OPTIONS":
        return {
            "statusCode": 200,
            "headers": {
                "Access-Control-Allow-Origin": "*",
                "Access-Control-Allow-Headers": "Content-Type",
                "Access-Control-Allow-Methods": "POST, OPTIONS",
            },
            "body": "",
        }

    try:
        body   = json.loads(event.get("body", "{}"))
        header  = body.get("header", {})
        entries = body.get("entries", [])

        # Path ke template (relative to function, template ada di public/templates)
        # Di Netlify, base dir = root project
        template_path = os.path.join(
            os.path.dirname(__file__), "..", "..", "public", "templates", "scenario-test-template.xlsx"
        )

        if not os.path.exists(template_path):
            return error_response("Template tidak ditemukan")

        wb = load_workbook(template_path)
        ws = wb.active

        # Sheet name
        safe = (header.get("judulDokumen") or "Sheet1").strip()[:31]
        ws.title = safe

        # Header fields
        ws["D4"] = header.get("keterangan", "")
        ws["D5"] = header.get("aplikasi", "")
        ws["D6"] = header.get("modul", "")
        ws["D7"] = header.get("createdBy", "")
        ws["D8"] = header.get("testedBy", "")
        ws["D9"] = header.get("targetFinish", "")
        ws["B12"] = header.get("judulDokumen", "")
        ws["B12"].font = Font(bold=True, size=13, name="Calibri")
        ws["B12"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Constants
        ROW_H_PT = 14.4
        COL_D_PX = 788
        ROW_H_PX = ROW_H_PT * 96 / 72  # ~19.2px

        thin = Side(style="thin", color="000000")

        def lr(cell):
            cell.border = Border(left=thin, right=thin)

        def lr_bottom(cell):
            cell.border = Border(left=thin, right=thin, bottom=thin)

        def f11(cell, bold=False, align="general", color="000000"):
            cell.font = Font(bold=bold, size=11, name="Calibri", color=color)
            cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)

        def status_font(status):
            if status == "NOT OK":
                return Font(bold=True, size=11, name="Calibri", color="FF0000")
            return Font(bold=(status == "OK"), size=11, name="Calibri", color="000000")

        def embed_image(data_url, start_row):
            if not data_url or "base64," not in data_url:
                return start_row
            try:
                b64 = data_url.split("base64,")[1]
                img_bytes = base64.b64decode(b64)
                pil = PILImage.open(io.BytesIO(img_bytes))
                orig_w, orig_h = pil.size
                disp_h = int(orig_h * COL_D_PX / orig_w)
                n_rows = max(1, int(disp_h / ROW_H_PX) + 1)

                for r in range(start_row, start_row + n_rows):
                    ws.row_dimensions[r].height = ROW_H_PT

                buf = io.BytesIO()
                pil.save(buf, format="PNG")
                buf.seek(0)
                xl_img = XLImage(buf)
                xl_img.width = COL_D_PX
                xl_img.height = disp_h
                ws.add_image(xl_img, f"D{start_row}")
                return start_row + n_rows
            except Exception:
                return start_row + 1

        current_row = 26

        for entry in entries:
            status = entry.get("status", "")
            tgl    = entry.get("tanggalTest", "")

            # Main entry row
            ws.row_dimensions[current_row].height = ROW_H_PT
            ws[f"B{current_row}"] = entry.get("no", "")
            ws[f"C{current_row}"] = entry.get("object", "")
            ws[f"D{current_row}"] = entry.get("keterangan", "")
            ws[f"E{current_row}"] = tgl
            ws[f"F{current_row}"] = status

            f11(ws[f"B{current_row}"], align="center")
            f11(ws[f"C{current_row}"], bold=True)
            ws[f"D{current_row}"].font = Font(size=11, name="Calibri")
            ws[f"D{current_row}"].alignment = Alignment(wrap_text=True, vertical="center")
            f11(ws[f"E{current_row}"], align="center")
            ws[f"F{current_row}"].font = status_font(status)
            ws[f"F{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

            for col in ["B", "C", "D", "E", "F"]:
                lr(ws[f"{col}{current_row}"])

            current_row += 1

            # Main images
            for img in entry.get("images", []):
                current_row = embed_image(img.get("dataUrl", ""), current_row)

            # Sub-sections
            for ss in entry.get("subSections", []):
                desc = ss.get("deskripsi", "")
                if desc:
                    ws.row_dimensions[current_row].height = ROW_H_PT
                    ws[f"D{current_row}"] = desc
                    ws[f"D{current_row}"].font = Font(size=11, name="Calibri")
                    ws[f"D{current_row}"].alignment = Alignment(wrap_text=True, vertical="center")
                    for col in ["B", "C", "D", "E", "F"]:
                        lr(ws[f"{col}{current_row}"])
                    current_row += 1

                for img in ss.get("images", []):
                    current_row = embed_image(img.get("dataUrl", ""), current_row)

            # Sub-keterangan legacy
            sub_k = entry.get("subKeterangan", "")
            if sub_k:
                ws.row_dimensions[current_row].height = ROW_H_PT
                ws[f"D{current_row}"] = sub_k
                ws[f"E{current_row}"] = tgl
                ws[f"D{current_row}"].font = Font(size=11, name="Calibri")
                ws[f"D{current_row}"].alignment = Alignment(vertical="center")
                ws[f"E{current_row}"].font = Font(size=11, name="Calibri")
                ws[f"E{current_row}"].alignment = Alignment(horizontal="center", vertical="center")
                for col in ["B", "C", "D", "E", "F"]:
                    lr(ws[f"{col}{current_row}"])
                current_row += 1

            # Bottom border di baris terakhir entri
            for col in ["B", "C", "D", "E", "F"]:
                lr_bottom(ws[f"{col}{current_row - 1}"])

        # Save to bytes
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        file_b64 = base64.b64encode(out.read()).decode("utf-8")

        fname = (header.get("judulDokumen") or "scenario-test").strip()
        for ch in r'\/:*?"<>|':
            fname = fname.replace(ch, "_")

        return {
            "statusCode": 200,
            "headers": {
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "Content-Disposition": f'attachment; filename="{fname}.xlsx"',
                "Content-Transfer-Encoding": "base64",
                "Access-Control-Allow-Origin": "*",
            },
            "body": file_b64,
            "isBase64Encoded": True,
        }

    except Exception as e:
        return error_response(str(e))


def error_response(msg):
    return {
        "statusCode": 500,
        "headers": {"Content-Type": "application/json", "Access-Control-Allow-Origin": "*"},
        "body": json.dumps({"error": msg}),
    }
